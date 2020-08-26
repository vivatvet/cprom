#!/usr/bin/env python3

from typing import Tuple
import math
import queue
import random
from openpyxl import load_workbook
from openpyxl import Workbook


class ExcelFile:

    def __init__(self):
        self.title = []

    @staticmethod
    def load_table_new(filepath: str) -> Tuple[list, list]:
        wb = load_workbook(filepath)
        sheet = wb.active
        title = []
        raw_table = []
        first_row = True
        for row in sheet:
            if first_row:
                for r in row:
                    title.append(str(r.value).strip())
                first_row = False
                continue
            row_a = [str(v.value).strip() for v in row]
            raw_table.append(row_a)
        return title, raw_table

    # def group_by_kd(self, raw_table: list):
    #     tb = {}
    #     for row in raw_table:
    #         tb.setdefault(row[5], []).append(row)
    #     for k in tb.keys():
    #         print(k + " -- " + str(len(tb[k])))

    def group_by_npp(self, raw_table: list) -> dict:
        tb = {}
        for rows in raw_table:
            work_rows = self.string_to_float_item(rows)
            tb.setdefault(rows[0], []).append(work_rows)
        return tb

    @staticmethod
    def string_to_float_item(rows: list) -> list:
        m_rows = []
        i = 0
        while i < len(rows):
            if i == 8:
                m_rows.append(float(rows[i].replace(',', '.')))
            else:
                m_rows.append(rows[i])
            i += 1
        return m_rows

    @staticmethod
    def sort_by_value(group_table: dict) -> dict:
        sorted_table = {}
        for group, rows in group_table.items():
            sorted_rows = sorted(rows, key=lambda x: x[8], reverse=True)
            sorted_table[group] = sorted_rows
        return sorted_table

    @staticmethod
    def get_90_cum_percent(group_table: dict) -> Tuple[dict, dict]:
        table = {}
        not_selected_table = {}
        for group, row in group_table.items():
            s = 0
            for r in row:
                s = s + r[8]
            previous_percent = 0
            percent = {}
            for i, r in enumerate(row):
                percent[i] = (r[8] / s) * 100
                if (previous_percent + percent[i] <= 90 or len(row) <= 3) \
                        or (i == 0 and percent[i] > 90) \
                        or (i == 1 and (percent[i] >= 10 and (percent[0] + percent[i] > 90))):
                    table.setdefault(group, []).append(r)
                else:
                    not_selected_table.setdefault(group, []).append(r)
                previous_percent = previous_percent + percent[i]
        return table, not_selected_table

    @staticmethod
    def get_stat_char(rows: list) -> Tuple[float, float, int, float]:
        s = 0
        for r in rows:
            s = s + r[8]
        average = s / len(rows)
        sig = 0
        for r in rows:
            sig = sig + (r[8] - average) * (r[8] - average)
        if len(rows) >= 20 or len(rows) == 1:
            sigma = math.sqrt(sig / (len(rows)))
        else:
            sigma = math.sqrt(sig / (len(rows) - 1))
        count = len(rows)
        covar = (sigma / average) * 100
        return average, sigma, count, covar

    @staticmethod
    def divide_rows(rows: list, average) -> Tuple[list, list]:
        rows_1 = []
        rows_2 = []
        for r in rows:
            if r[8] > average:
                rows_1.append(r)
            else:
                rows_2.append(r)
        return rows_1, rows_2

    def stat_char(self, group_table: dict) -> Tuple[dict, dict, dict, dict, dict, dict]:
        averages = {}
        sigmas = {}
        counts = {}
        covars = {}
        chosen = {}
        strata = {}
        not_chosen = {}
        q = queue.Queue()
        for group, rows in group_table.items():
            averages[group] = {}
            sigmas[group] = {}
            counts[group] = {}
            step = 0
            averages[group][step], sigmas[group][step], counts[group][step], _ = self.get_stat_char(rows)
            for r in rows:
                if len(rows) >= 20:
                    if r[8] >= (averages[group][step] + (3 * sigmas[group][step])):
                        chosen.setdefault(group, []).append(r)
                    else:
                        not_chosen.setdefault(group, []).append(r)
                elif len(rows) <= 3:
                    chosen.setdefault(group, []).append(r)
                else:
                    if r[8] >= (averages[group][step] + (2 * sigmas[group][step])):
                        chosen.setdefault(group, []).append(r)
                    else:
                        not_chosen.setdefault(group, []).append(r)
        for group, rows in not_chosen.items():
            q.put(rows)
            while not q.empty():
                rows_q = q.get()
                average, sigma, count, covar = self.get_stat_char(rows_q)
                if covar > 33:
                    row_1, row_2 = self.divide_rows(rows_q, average)
                    q.put(row_1)
                    q.put(row_2)
                else:
                    strata.setdefault(group, []).append(rows_q)
        return chosen, strata, averages, sigmas, counts, covars

    def random_choose(self, strata: dict, average: dict) -> dict:
        random_chosen = {}
        for group, strata_item in strata.items():
            dn = 0
            count_group = 0
            sum_group = 0
            for i, rows in enumerate(strata_item):
                _, sigma, count, _ = self.get_stat_char(rows)
                dn = dn + (sigma * sigma * count)
                count_group = count_group + count
                for r in rows:
                    sum_group = sum_group + r[8]
            variance_group = dn / count_group
            opt_number = (count_group * variance_group * (1.96 * 1.96)) / ((variance_group * (1.96 * 1.96)) + ((0.1 * average[group][0]) * (0.1 * average[group][0]) * count_group))
            for i, rows in enumerate(strata_item):
                sum_strata = 0
                for r in rows:
                    sum_strata = sum_strata + r[8]
                opt_number_strata = math.ceil((sum_strata / sum_group) * (math.ceil(opt_number + 2)))
                rows_dict = {k: v for k, v in enumerate(rows)}
                choose_number = opt_number_strata if len(rows_dict) >= opt_number_strata else len(rows_dict)
                choose_list = random.sample(rows_dict.keys(), choose_number)
                choose_list.sort()
                rows_for_choose = [rows_dict[k] + ['chosen'] if k in choose_list else rows_dict[k] for k in rows_dict.keys()]
                random_chosen.setdefault(group, []).append(rows_for_choose)
        return random_chosen

    def make_final_table(self, tb_by_npp: dict, chosen: dict, strata: dict, not_selected: dict) -> dict:
        final_table = {}
        for group in tb_by_npp.keys():
            if chosen.get(group):
                for rows in chosen[group]:
                    final_table.setdefault(group, {}).setdefault('big', []).append(rows + ['BIG'])
            if strata.get(group):
                i = 1
                for strata_item in strata[group]:
                    strata_name = 'strata_' + str(i)
                    average, sigma, _, covar = self.get_stat_char(strata_item)
                    first = True
                    for row in strata_item:
                        if first:
                            if row[-1] == 'chosen':
                                r = row + [strata_name, average, sigma, covar]
                            else:
                                r = row + ['', strata_name, average, sigma, covar]
                        else:
                            if row[-1] == 'chosen':
                                r = row + [strata_name]
                            else:
                                r = row + ['', strata_name]
                        final_table.setdefault(group, {}).setdefault(strata_name, []).append(r)
                        first = False
                    i += 1
            if not_selected.get(group):
                for rows in not_selected[group]:
                    final_table.setdefault(group, {}).setdefault('not_selected', []).append(rows)
        return final_table

    def is_file_exist(self, file_name):
        pass

    @staticmethod
    def get_sum_all(table: dict) -> float:
        sum_all = 0
        for _, rows in table.items():
            for row in rows:
                sum_all = sum_all + row[8]
        return sum_all

    @staticmethod
    def get_sum_chosen(table: dict) -> float:
        sum_chosen = 0
        for _, rows in table.items():
            for row in rows:
                if 'BIG' in row or 'chosen' in row:
                    sum_chosen = sum_chosen + row[8]
        return sum_chosen

    def write_to_file_new(self, file_name: str, tb_title: list, final_table: dict):
        wb = Workbook()
        wb.save(file_name)
        wb = load_workbook(file_name)
        sheet = wb.active

        for col_id, data in enumerate(tb_title + ['', 'strata', 'average', 'sigma', 'covar'], start=1):
            sheet.cell(row=1, column=col_id).value = data
        i = 2
        for group in final_table.keys():
            sum_all = self.get_sum_all(final_table[group])
            sum_chosen = self.get_sum_chosen(final_table[group])
            percent = (sum_chosen / sum_all) * 100
            for col_id, data in enumerate([group, sum_all, percent], start=1):
                sheet.cell(row=i, column=col_id).value = data
            i += 1
            # write big
            if final_table[group].get('big'):
                for rows in final_table[group]['big']:
                    for col_id, data in enumerate(rows, start=1):
                        sheet.cell(row=i, column=col_id).value = data
                    i += 1
            # write stratas
            i_strata = 1
            is_strata = True
            while is_strata:
                if final_table[group].get('strata_' + str(i_strata)):
                    for rows in final_table[group]['strata_' + str(i_strata)]:
                        for col_id, data in enumerate(rows, start=1):
                            sheet.cell(row=i, column=col_id).value = data
                        i += 1
                    i_strata += 1
                else:
                    is_strata = False
            # write not selected
            if final_table[group].get('not_selected'):
                for rows in final_table[group]['not_selected']:
                    for col_id, data in enumerate(rows, start=1):
                        sheet.cell(row=i, column=col_id).value = data
                    i += 1
            i += 1
        wb.save(file_name)
